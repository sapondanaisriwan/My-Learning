from selenium import webdriver
from openpyxl import load_workbook

driver = webdriver.Edge("msedgedriver.exe")


def login_to_web(url='https://www.noraphat.dev/POS_NAME/index.php'):
    driver.get(url)

    username = driver.find_element_by_name('username')
    username.send_keys('admin')

    password = driver.find_element_by_name('password')
    password.send_keys('1234')

    # driver.find_element_by_xpath("/html/body/div/div/div/form/div[3]/div[2]/button")
    # driver.find_element_by_name('btnLogin')
    press_button = driver.find_element_by_name('btnLogin')
    press_button.click()


def add_product(url, exc_product_name, exc_product_quantity, exc_product_price):
    driver.get(url)

    pronduct_name = driver.find_element_by_id('product_name')
    pronduct_name.send_keys(exc_product_name)

    amount_product = driver.find_element_by_name('product_quantity')
    amount_product.send_keys(exc_product_quantity)

    product_price = driver.find_element_by_name('product_price')
    product_price.send_keys(exc_product_price)

    product_image = driver.find_element_by_name('post_image')
    product_image.send_keys("D:\My Photo\Meme\คือลือ.png")

    # submit = driver.find_element_by_name('btnAdd')
    # submit.click()
    # driver.find_element_by_xpath("/html/body/div/div/div/form/div[3]/div[2]/button")
    # driver.find_element_by_name('btnLogin')
    press_button = driver.find_element_by_name('btnLogin')
    press_button.click()


login_to_web()

excelfile = load_workbook('databag.xlsx')
sheet = excelfile['DataBag']
maxrow = int(sheet.max_row)

for i in range(2, maxrow + 1):
    exc_product_name = sheet.cell(row=i, column=1).value
    exc_product_quantity = sheet.cell(row=i, column=2).value
    exc_product_price = sheet.cell(row=i, column=3).value
    add_product('https://www.noraphat.dev/POS_NAME/add-product.php', exc_product_name, exc_product_quantity, exc_product_price)
