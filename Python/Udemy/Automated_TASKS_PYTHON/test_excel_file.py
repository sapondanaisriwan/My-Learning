from re import S
from openpyxl import load_workbook
from openpyxl.reader import excel

excelfile = load_workbook('databag.xlsx')
sheet = excelfile['DataBag']
maxrow = int(sheet.max_row)

for i in range(2, maxrow + 1):
    print(i)
    txt_product_name = sheet.cell(row=i, column=1).value
    txt_product_quantity = sheet.cell(row=i, column=2).value
    txt_product_price = sheet.cell(row=i, column=3).value
    print(txt_product_name, txt_product_quantity, txt_product_price)